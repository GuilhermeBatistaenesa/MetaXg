import json
import logging
import os
import uuid
import getpass
import platform
from datetime import datetime, date, time
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

DEFAULT_PATH = r"P:\AuditoriaRobos\Auditoria_Robos.xlsx"
PENDING_DIR = r"P:\AuditoriaRobos\pending"
ROBO_NOME = "MetaX"
ORIGEM_CODIGO = r"P:\ProcessosMetaX\codigos"
VERSION_FALLBACK = "1.0.0"

RUNS_HEADERS = [
    "run_id",
    "robo_nome",
    "versao_robo",
    "ambiente",
    "data_execucao",
    "hora_inicio",
    "hora_fim",
    "duracao_segundos",
    "total_processado",
    "total_sucesso",
    "total_erro",
    "taxa_sucesso",
    "taxa_erro",
    "erros_auto_mitigados",
    "erros_manuais",
    "erros_pendentes",
    "resultado_final",
    "observacoes",
    "host_maquina",
    "usuario_execucao",
    "origem_codigo",
    "commit_hash",
    "build_id",
]

ERRORS_HEADERS = [
    "run_id",
    "robo_nome",
    "timestamp",
    "etapa",
    "tipo_erro",
    "codigo_erro",
    "mensagem_resumida",
    "registro_id",
    "mitigacao",
    "resolvido_em",
]

ROBOS_HEADERS = [
    "robo_nome",
    "origem_codigo",
    "versao_robo",
    "ambiente_padrao",
    "observacoes",
]

LIST_AMBIENTE = ["PROD", "TEST"]
LIST_TIPO_ERRO = [
    "Tecnico",
    "Regra de Negocio",
    "Dados Invalidos",
    "Externo/Indisponibilidade",
]
LIST_MITIGACAO = ["Auto", "Manual", "Pendente"]
LIST_RESULTADO_FINAL = ["100% Concluído", "Concluído com Exceções", "Incompleto"]

_logger = logging.getLogger("auditoria_excel")
if not _logger.handlers:
    logging.basicConfig(level=logging.INFO, format="[auditoria_excel] %(levelname)s: %(message)s")


def log_run(run_data: dict, errors: list[dict] | None = None, path: str = DEFAULT_PATH) -> dict:
    if run_data is None:
        run_data = {}
    errors = errors or []

    run_id = str(run_data.get("run_id") or uuid.uuid4())
    robo_nome = str(run_data.get("robo_nome") or ROBO_NOME)
    versao_robo = str(run_data.get("versao_robo") or _read_version())
    ambiente = str(run_data.get("ambiente") or os.getenv("METAX_ENV", "PROD")).upper()
    if ambiente not in LIST_AMBIENTE:
        ambiente = "PROD"

    started_at = _coerce_datetime(run_data.get("started_at")) or datetime.now()
    finished_at = _coerce_datetime(run_data.get("finished_at")) or datetime.now()
    duration_sec = run_data.get("duration_sec")
    if duration_sec is None:
        duration_sec = int((finished_at - started_at).total_seconds())
    try:
        duration_sec = int(duration_sec)
    except Exception:
        duration_sec = 0
    if duration_sec < 0:
        duration_sec = 0

    total_processado = _safe_int(run_data.get("total_processado"))
    total_sucesso = _safe_int(run_data.get("total_sucesso"))
    total_erro = _safe_int(run_data.get("total_erro"))
    erros_auto = _safe_int(run_data.get("erros_auto_mitigados"))
    erros_manuais = _safe_int(run_data.get("erros_manuais"))

    observacoes = str(run_data.get("observacoes") or "")
    host_maquina = str(run_data.get("host_maquina") or platform.node())
    usuario_execucao = str(run_data.get("usuario_execucao") or getpass.getuser())
    origem_codigo = str(run_data.get("origem_codigo") or ORIGEM_CODIGO)
    commit_hash = str(run_data.get("commit_hash") or _env_first(["GIT_COMMIT", "GITHUB_SHA", "CI_COMMIT_SHA"]) or "")
    build_id = str(run_data.get("build_id") or _env_first(["BUILD_ID", "GITHUB_RUN_ID", "CI_PIPELINE_ID"]) or "")

    resultado_final = run_data.get("resultado_final")
    if not resultado_final:
        pendentes = max(total_erro - erros_auto - erros_manuais, 0)
        if total_processado == 0 or (total_sucesso + total_erro) != total_processado:
            resultado_final = "Incompleto"
        elif total_erro == (erros_auto + erros_manuais) and pendentes == 0:
            resultado_final = "100% Concluído"
        else:
            resultado_final = "Concluído com Exceções"

    run_row = {
        "run_id": run_id,
        "robo_nome": robo_nome,
        "versao_robo": versao_robo,
        "ambiente": ambiente,
        "data_execucao": started_at.date(),
        "hora_inicio": started_at.time(),
        "hora_fim": finished_at.time(),
        "duracao_segundos": duration_sec,
        "total_processado": total_processado,
        "total_sucesso": total_sucesso,
        "total_erro": total_erro,
        "erros_auto_mitigados": erros_auto,
        "erros_manuais": erros_manuais,
        "resultado_final": resultado_final,
        "observacoes": observacoes,
        "host_maquina": host_maquina,
        "usuario_execucao": usuario_execucao,
        "origem_codigo": origem_codigo,
        "commit_hash": commit_hash,
        "build_id": build_id,
    }

    wb = _load_or_create_workbook(path)
    runs_ws = wb["RUNS"]
    errors_ws = wb["ERRORS"]
    robos_ws = wb["ROBOS"]

    _ensure_validations_runs(runs_ws)
    _ensure_validations_errors(errors_ws)

    run_row_index = _append_run_row(runs_ws, run_row)
    _append_errors(errors_ws, errors, run_id=run_id, robo_nome=robo_nome)
    _ensure_robot_registry(robos_ws, robo_nome, origem_codigo, versao_robo, ambiente)

    _update_table_range(runs_ws, "RunsTable", run_row_index)
    _update_table_range(errors_ws, "ErrorsTable", errors_ws.max_row)
    _update_table_range(robos_ws, "RobosTable", robos_ws.max_row)

    try:
        wb.save(path)
        _logger.info(f"Auditoria salva em {path}")
        return {
            "run_id": run_id,
            "saved_path": path,
            "fallback": False,
        }
    except Exception as exc:
        _logger.warning(f"Falha ao salvar auditoria principal: {exc}")
        pending_path = _build_pending_path()
        _safe_mkdir(PENDING_DIR)
        saved_path = None
        try:
            wb.save(pending_path)
            saved_path = pending_path
            _logger.info(f"Auditoria salva em fallback: {pending_path}")
        except Exception as fallback_exc:
            _logger.error(f"Falha ao salvar auditoria fallback: {fallback_exc}")
        _write_pending_json(run_id, run_row, errors, path, saved_path, str(exc))
        return {
            "run_id": run_id,
            "saved_path": saved_path,
            "fallback": True,
            "error": str(exc),
        }


def _append_run_row(ws, row_data: dict) -> int:
    row_index = _next_empty_row(ws, key_col=1, start_row=2)
    if row_index < 2:
        row_index = 2

    for col_index, header in enumerate(RUNS_HEADERS, start=1):
        cell = ws.cell(row=row_index, column=col_index)
        value = row_data.get(header)
        if header == "taxa_sucesso":
            value = _formula_taxa_sucesso(row_index)
        elif header == "taxa_erro":
            value = _formula_taxa_erro(row_index)
        elif header == "erros_pendentes":
            value = _formula_erros_pendentes(row_index)
        cell.value = value

    _apply_run_row_formats(ws, row_index)
    return row_index


def _append_errors(ws, errors: Iterable[dict], run_id: str, robo_nome: str):
    if not errors:
        return
    for error in errors:
        row_index = _next_empty_row(ws, key_col=1, start_row=2)
        normalized = _normalize_error_dict(error, run_id, robo_nome)
        for col_index, header in enumerate(ERRORS_HEADERS, start=1):
            ws.cell(row=row_index, column=col_index).value = normalized.get(header)
        _apply_error_row_formats(ws, row_index)


def _ensure_robot_registry(ws, robo_nome: str, origem_codigo: str, versao_robo: str, ambiente: str):
    existing = _find_row_by_value(ws, 1, robo_nome)
    if existing:
        return
    row_index = _next_empty_row(ws, key_col=1, start_row=2)
    values = [
        robo_nome,
        origem_codigo,
        versao_robo,
        ambiente,
        "",
    ]
    for col_index, value in enumerate(values, start=1):
        ws.cell(row=row_index, column=col_index).value = value


def _find_row_by_value(ws, col_index: int, value: str) -> int | None:
    if not value:
        return None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=col_index).value or "") == str(value):
            return row
    return None


def _apply_run_row_formats(ws, row_index: int):
    date_col = RUNS_HEADERS.index("data_execucao") + 1
    time_start_col = RUNS_HEADERS.index("hora_inicio") + 1
    time_end_col = RUNS_HEADERS.index("hora_fim") + 1
    taxa_sucesso_col = RUNS_HEADERS.index("taxa_sucesso") + 1
    taxa_erro_col = RUNS_HEADERS.index("taxa_erro") + 1
    ws.cell(row=row_index, column=date_col).number_format = "yyyy-mm-dd"
    ws.cell(row=row_index, column=time_start_col).number_format = "hh:mm:ss"
    ws.cell(row=row_index, column=time_end_col).number_format = "hh:mm:ss"
    ws.cell(row=row_index, column=taxa_sucesso_col).number_format = "0.00%"
    ws.cell(row=row_index, column=taxa_erro_col).number_format = "0.00%"


def _apply_error_row_formats(ws, row_index: int):
    ts_col = ERRORS_HEADERS.index("timestamp") + 1
    resolvido_col = ERRORS_HEADERS.index("resolvido_em") + 1
    ws.cell(row=row_index, column=ts_col).number_format = "yyyy-mm-dd hh:mm:ss"
    ws.cell(row=row_index, column=resolvido_col).number_format = "yyyy-mm-dd hh:mm:ss"


def _formula_taxa_sucesso(row_index: int) -> str:
    return f"=IF($I{row_index}>0,$J{row_index}/$I{row_index},0)"


def _formula_taxa_erro(row_index: int) -> str:
    return f"=IF($I{row_index}>0,$K{row_index}/$I{row_index},0)"


def _formula_erros_pendentes(row_index: int) -> str:
    return f"=MAX($K{row_index}-$N{row_index}-$O{row_index},0)"


def _normalize_error_dict(error: dict, run_id: str, robo_nome: str) -> dict:
    if error is None:
        error = {}
    timestamp = _coerce_datetime(error.get("timestamp")) or datetime.now()
    resolvido_em = _coerce_datetime(error.get("resolvido_em"))

    return {
        "run_id": error.get("run_id") or run_id,
        "robo_nome": error.get("robo_nome") or robo_nome,
        "timestamp": timestamp,
        "etapa": str(error.get("etapa") or ""),
        "tipo_erro": str(error.get("tipo_erro") or "Tecnico"),
        "codigo_erro": str(error.get("codigo_erro") or ""),
        "mensagem_resumida": str(error.get("mensagem_resumida") or "")[:250],
        "registro_id": str(error.get("registro_id") or ""),
        "mitigacao": str(error.get("mitigacao") or "Pendente"),
        "resolvido_em": resolvido_em,
    }


def _load_or_create_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        wb = load_workbook(path)
        _ensure_sheets(wb)
        return wb
    wb = Workbook()
    _create_template(wb)
    return wb


def _ensure_sheets(wb: Workbook):
    if "RUNS" not in wb.sheetnames:
        ws = wb.create_sheet("RUNS")
        _setup_runs_sheet(ws)
    if "ERRORS" not in wb.sheetnames:
        ws = wb.create_sheet("ERRORS")
        _setup_errors_sheet(ws)
    if "ROBOS" not in wb.sheetnames:
        ws = wb.create_sheet("ROBOS")
        _setup_robos_sheet(ws)
    if "DASHBOARD" not in wb.sheetnames:
        ws = wb.create_sheet("DASHBOARD")
        _setup_dashboard_sheet(wb, ws)


def _create_template(wb: Workbook):
    if wb.active:
        wb.remove(wb.active)

    ws_runs = wb.create_sheet("RUNS")
    ws_errors = wb.create_sheet("ERRORS")
    ws_robos = wb.create_sheet("ROBOS")
    ws_dash = wb.create_sheet("DASHBOARD")

    _setup_runs_sheet(ws_runs)
    _setup_errors_sheet(ws_errors)
    _setup_robos_sheet(ws_robos)
    _setup_dashboard_sheet(wb, ws_dash)


def _setup_runs_sheet(ws):
    ws.append(RUNS_HEADERS)
    ws.append([None] * len(RUNS_HEADERS))
    _style_header_row(ws, len(RUNS_HEADERS))
    ws.freeze_panes = "A2"

    _set_column_widths(ws, {
        "A": 38,
        "B": 14,
        "C": 12,
        "D": 10,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 16,
        "I": 16,
        "J": 14,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 16,
        "O": 14,
        "P": 14,
        "Q": 20,
        "R": 28,
        "S": 18,
        "T": 18,
        "U": 28,
        "V": 16,
        "W": 16,
    })

    table_ref = f"A1:{get_column_letter(len(RUNS_HEADERS))}2"
    table = Table(displayName="RunsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    _ensure_validations_runs(ws)
    _apply_conditional_formatting_runs(ws)


def _setup_errors_sheet(ws):
    ws.append(ERRORS_HEADERS)
    ws.append([None] * len(ERRORS_HEADERS))
    _style_header_row(ws, len(ERRORS_HEADERS))
    ws.freeze_panes = "A2"

    _set_column_widths(ws, {
        "A": 38,
        "B": 14,
        "C": 20,
        "D": 20,
        "E": 22,
        "F": 16,
        "G": 60,
        "H": 18,
        "I": 14,
        "J": 20,
    })

    table_ref = f"A1:{get_column_letter(len(ERRORS_HEADERS))}2"
    table = Table(displayName="ErrorsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    _ensure_validations_errors(ws)


def _setup_robos_sheet(ws):
    ws.append(ROBOS_HEADERS)
    ws.append([None] * len(ROBOS_HEADERS))
    _style_header_row(ws, len(ROBOS_HEADERS))
    ws.freeze_panes = "A2"
    _set_column_widths(ws, {
        "A": 20,
        "B": 40,
        "C": 14,
        "D": 16,
        "E": 30,
    })

    table_ref = f"A1:{get_column_letter(len(ROBOS_HEADERS))}2"
    table = Table(displayName="RobosTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _setup_dashboard_sheet(wb: Workbook, ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    kpi_label_font = Font(name="Segoe UI", size=11, bold=True, color="1F2933")
    kpi_value_font = Font(name="Segoe UI", size=14, bold=True, color="111827")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    card_fill = PatternFill("solid", fgColor="E6EEF7")

    ws["A1"].value = "AUDITORIA ROBOS - DASHBOARD"
    ws["A1"].font = title_font
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:F1")
    for col in range(1, 7):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("Execucoes 7 dias", "=COUNTIFS(RUNS!E:E,\">=\"&TODAY()-6,RUNS!E:E,\"<=\"&TODAY())"),
        ("Execucoes 30 dias", "=COUNTIFS(RUNS!E:E,\">=\"&TODAY()-29,RUNS!E:E,\"<=\"&TODAY())"),
        ("Taxa media sucesso", "=IFERROR(AVERAGE(RUNS!L:L),0)"),
        ("Total erros", "=SUM(RUNS!K:K)"),
        ("% mitigado auto", "=IFERROR(SUM(RUNS!N:N)/SUM(RUNS!K:K),0)"),
        ("Tempo medio execucao (s)", "=IFERROR(AVERAGE(RUNS!H:H),0)"),
    ]

    start_row = 3
    for idx, (label, formula) in enumerate(kpis, start=0):
        row = start_row + idx
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = kpi_label_font
        ws[f"A{row}"].fill = card_fill
        ws[f"B{row}"].value = formula
        ws[f"B{row}"].font = kpi_value_font
        ws[f"B{row}"].fill = card_fill
        ws[f"A{row}"].alignment = Alignment(horizontal="left")
        ws[f"B{row}"].alignment = Alignment(horizontal="right")
        ws.row_dimensions[row].height = 20

    ws["B5"].number_format = "0.00%"
    ws["B7"].number_format = "0.00%"

    _setup_dashboard_charts(wb, ws)


def _setup_dashboard_charts(wb: Workbook, ws):
    runs_ws = wb["RUNS"]
    errors_ws = wb["ERRORS"]

    max_rows = 5000

    line_chart = LineChart()
    line_chart.title = "Taxa de sucesso no tempo"
    line_chart.y_axis.title = "Taxa"
    line_chart.x_axis.title = "Data"
    data = Reference(runs_ws, min_col=12, min_row=1, max_row=max_rows)
    dates = Reference(runs_ws, min_col=5, min_row=2, max_row=max_rows)
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(dates)
    line_chart.height = 7
    line_chart.width = 16
    ws.add_chart(line_chart, "D3")

    bar_chart = BarChart()
    bar_chart.title = "Processado vs Erro"
    bar_chart.y_axis.title = "Volume"
    bar_chart.x_axis.title = "Execucoes"
    data = Reference(runs_ws, min_col=9, max_col=11, min_row=1, max_row=max_rows)
    cats = Reference(runs_ws, min_col=5, min_row=2, max_row=max_rows)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.height = 7
    bar_chart.width = 16
    ws.add_chart(bar_chart, "D18")

    ws["A12"].value = "Tipo_erro"
    ws["B12"].value = "Qtde"
    _style_header_row(ws, 2, row=12)
    for idx, tipo in enumerate(LIST_TIPO_ERRO, start=13):
        ws[f"A{idx}"].value = tipo
        ws[f"B{idx}"].value = f"=COUNTIF(ERRORS!E:E,A{idx})"

    pie_chart = PieChart()
    pie_chart.title = "Erros por tipo"
    data = Reference(ws, min_col=2, min_row=12, max_row=12 + len(LIST_TIPO_ERRO))
    labels = Reference(ws, min_col=1, min_row=13, max_row=12 + len(LIST_TIPO_ERRO))
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.height = 7
    pie_chart.width = 12
    ws.add_chart(pie_chart, "D33")

    dur_chart = LineChart()
    dur_chart.title = "Duracao por execucao"
    dur_chart.y_axis.title = "Segundos"
    dur_chart.x_axis.title = "Data"
    data = Reference(runs_ws, min_col=8, min_row=1, max_row=max_rows)
    cats = Reference(runs_ws, min_col=5, min_row=2, max_row=max_rows)
    dur_chart.add_data(data, titles_from_data=True)
    dur_chart.set_categories(cats)
    dur_chart.height = 7
    dur_chart.width = 16
    ws.add_chart(dur_chart, "D48")


def _ensure_validations_runs(ws):
    _add_list_validation(ws, "D2:D1048576", LIST_AMBIENTE)
    _add_list_validation(ws, "Q2:Q1048576", LIST_RESULTADO_FINAL)


def _ensure_validations_errors(ws):
    _add_list_validation(ws, "E2:E1048576", LIST_TIPO_ERRO)
    _add_list_validation(ws, "I2:I1048576", LIST_MITIGACAO)


def _apply_conditional_formatting_runs(ws):
    red_fill = PatternFill("solid", fgColor="F8D7DA")
    green_fill = PatternFill("solid", fgColor="D4EDDA")

    ws.conditional_formatting.add(
        "A2:W1048576",
        FormulaRule(formula=["=$J2+$K2<>$I2"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        "M2:M1048576",
        CellIsRule(operator="greaterThan", formula=["0.05"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        "Q2:Q1048576",
        FormulaRule(formula=["=EXACT($Q2,\"100% Concluído\")"], fill=green_fill),
    )


def _add_list_validation(ws, cell_range: str, items: list[str]):
    if not items:
        return
    formula = '"' + ",".join(items) + '"'
    existing = getattr(ws.data_validations, "dataValidation", [])
    for dv in existing:
        if dv.formula1 == formula and cell_range in str(dv.ranges):
            return
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _style_header_row(ws, total_cols: int, row: int = 1):
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border


def _set_column_widths(ws, widths: dict[str, int]):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _next_empty_row(ws, key_col: int, start_row: int = 2) -> int:
    row = start_row
    while row <= ws.max_row:
        value = ws.cell(row=row, column=key_col).value
        if value is None or str(value).strip() == "":
            return row
        row += 1
    return ws.max_row + 1


def _update_table_range(ws, table_name: str, last_row: int):
    if table_name not in ws.tables:
        return
    table = ws.tables[table_name]
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    if last_row < min_row + 1:
        last_row = min_row + 1
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_row}"


def _build_pending_path() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(PENDING_DIR, f"Auditoria_Robos__PENDENTE__{timestamp}.xlsx")


def _write_pending_json(run_id: str, run_row: dict, errors: list[dict], target_path: str, saved_path: str | None, error: str):
    payload = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(),
        "target_path": target_path,
        "saved_path": saved_path,
        "error": error,
        "run": _json_safe(run_row),
        "errors": [_json_safe(e) for e in (errors or [])],
    }
    _safe_mkdir(PENDING_DIR)
    json_path = os.path.join(PENDING_DIR, f"run__{run_id}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def _safe_mkdir(path: str):
    try:
        os.makedirs(path, exist_ok=True)
    except Exception:
        pass


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    return value


def _coerce_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue
    return None


def _safe_int(value: Any) -> int:
    try:
        return int(value)
    except Exception:
        return 0


def _read_version() -> str:
    version_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "version.txt")
    if os.path.exists(version_path):
        try:
            with open(version_path, "r", encoding="utf-8") as f:
                raw = f.read().strip()
            if raw:
                return raw
        except Exception:
            pass
    return VERSION_FALLBACK


def _env_first(keys: list[str]) -> str | None:
    for key in keys:
        value = os.getenv(key)
        if value:
            return value
    return None
